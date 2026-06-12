import os
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# Import core functions from our pipeline
from pipeline import (
    load_and_clean_data,
    forecast_all_products_parallel,
    format_forecast_table,
    forecast_single_product,
    prepare_product_series
)

# Set up Streamlit Page Page Layout
st.set_page_config(
    page_title="Prophet Sales Forecast UI",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inject custom styling for modern look and premium colors
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1E3A8A;
        margin-bottom: 0.2rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #4B5563;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #F3F4F6;
        border-radius: 8px;
        padding: 15px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        text-align: center;
    }
    .metric-val {
        font-size: 1.8rem;
        font-weight: bold;
        color: #111827;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #6B7280;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">Product-Wise 10-Day Sales Forecast</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Upload historical transactional data to forecast individual daily product sales using Prophet.</div>', unsafe_allow_html=True)

# ----------------- Sidebar -----------------
st.sidebar.header("📁 Data Source & Config")

uploaded_file = st.sidebar.file_uploader(
    "Upload Sales Data (.xlsx or .csv)", 
    type=["xlsx", "csv"],
    help="Required columns: SlNo, INVOICEDATE, ITEMID, QTY"
)

st.sidebar.markdown("---")
st.sidebar.header("⏱️ Forecast Configuration")

forecast_unit = st.sidebar.radio("Forecast Period Unit", ["Days", "Weeks"])

if forecast_unit == "Days":
    forecast_periods = st.sidebar.slider("Forecast Duration (Days)", min_value=1, max_value=10, value=10)
    default_high = 50
    default_med = 10
else:
    forecast_periods = st.sidebar.slider("Forecast Duration (Weeks)", min_value=1, max_value=5, value=4)
    default_high = 350
    default_med = 70

st.sidebar.markdown("---")
st.sidebar.header("🎨 Highlighting Thresholds")
st.sidebar.info("Highlight forecasted quantities sold per day/week using colors:")

high_thresh = st.sidebar.number_input("High Volume (Green) Threshold", min_value=1, value=default_high, step=5)
med_thresh = st.sidebar.number_input("Medium Volume (Orange) Threshold", min_value=0, value=default_med, step=1)

# Helper function to generate styled Excel for download
def export_to_excel_styled(forecast_df, date_cols, med, high):
    wb = Workbook()
    ws = wb.active
    ws.title = "10-Day Forecast"
    
    # Fonts and Fills
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    green_fill = PatternFill(start_color="EAF3DE", end_color="EAF3DE", fill_type="solid")
    green_font = Font(name="Segoe UI", size=11, color="27500A", bold=True)
    
    orange_fill = PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid")
    orange_font = Font(name="Segoe UI", size=11, color="633806", bold=True)
    
    red_fill = PatternFill(start_color="FCEBEB", end_color="FCEBEB", fill_type="solid")
    red_font = Font(name="Segoe UI", size=11, color="791F1F")
    
    default_font = Font(name="Segoe UI", size=11)
    
    # Write Headers
    headers = list(forecast_df.columns)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        
    # Write Rows and Apply Styles
    for row_idx, row_data in enumerate(forecast_df.values, start=2):
        row_list = list(row_data)
        ws.append(row_list)
        
        # Color coding forecast cells
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = default_font
            
            if col_name in date_cols:
                val = cell.value
                if val is not None:
                    try:
                        val = float(val)
                        if val >= high:
                            cell.fill = green_fill
                            cell.font = green_font
                        elif val >= med:
                            cell.fill = orange_fill
                            cell.font = orange_font
                        else:
                            cell.fill = red_fill
                            cell.font = red_font
                    except ValueError:
                        pass
                        
    # Freeze pane (top row and first column)
    ws.freeze_panes = "B2"
    
    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Output to bytes buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# ----------------- Main UI Controller -----------------
if uploaded_file is not None:
    # 1. Load Data
    with st.spinner("Parsing and cleaning dataset..."):
        try:
            # We save the file to a temporary location or pass it directly if we read bytes.
            # pipeline.load_and_clean_data accepts a file path, so we write the uploaded file to a local temp file.
            temp_path = os.path.join("output", "temp_upload.xlsx")
            os.makedirs("output", exist_ok=True)
            with open(temp_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
                
            df_cleaned = load_and_clean_data(temp_path)
        except Exception as e:
            st.error(f"⚠️ {e}")
            st.stop()

    # 2. Display summary metrics
    col1, col2, col3, col4 = st.columns(4)
    
    unique_items = sorted(df_cleaned['ITEMID'].dropna().unique())
    min_date = df_cleaned['INVOICEDATE'].min().strftime('%Y-%m-%d')
    max_date = df_cleaned['INVOICEDATE'].max().strftime('%Y-%m-%d')
    total_qty = df_cleaned['QTY'].sum()
    
    col1.markdown(f'<div class="metric-card"><div class="metric-val">{len(unique_items)}</div><div class="metric-label">Total Products</div></div>', unsafe_allow_html=True)
    col2.markdown(f'<div class="metric-card"><div class="metric-val">{min_date}</div><div class="metric-label">Start Date</div></div>', unsafe_allow_html=True)
    col3.markdown(f'<div class="metric-card"><div class="metric-val">{max_date}</div><div class="metric-label">End Date</div></div>', unsafe_allow_html=True)
    col4.markdown(f'<div class="metric-card"><div class="metric-val">{total_qty:,}</div><div class="metric-label">Historical Units Sold</div></div>', unsafe_allow_html=True)
    
    st.markdown("---")
    
    # 3. Action panel
    st.subheader("🔮 Run Forecast")
    st.markdown(f"Fit individual Prophet forecasting models for each product to predict sales over the next {forecast_periods} {forecast_unit.lower()}.")
    
    button_label = f"🚀 Generate {forecast_periods}-{forecast_unit[:-1]} Product-Wise Forecast"
    if st.button(button_label):
        progress_bar = st.progress(0)
        progress_text = st.empty()
        
        # Callback to update streamlit progress bar
        def progress_cb(completed, total):
            progress_bar.progress(completed / total)
            progress_text.text(f"Processed {completed} of {total} products...")
            
        # Determine total forecast days
        days_to_predict = forecast_periods if forecast_unit == "Days" else forecast_periods * 7
            
        with st.spinner("Fitting Prophet models in parallel..."):
            forecast_dict = forecast_all_products_parallel(
                df_cleaned, 
                forecast_days=days_to_predict, 
                progress_callback=progress_cb
            )
            
        progress_bar.empty()
        progress_text.empty()
        
        # Save results in session state to prevent losing them on interactions
        st.session_state['forecast_dict'] = forecast_dict
        st.session_state['forecast_table'] = format_forecast_table(forecast_dict, forecast_unit)
        st.session_state['forecast_unit_used'] = forecast_unit
        st.session_state['forecast_periods_used'] = forecast_periods
        st.success("Successfully completed forecasting for all products!")

    # 4. Display Forecast Table if generated
    if 'forecast_table' in st.session_state:
        st.markdown("---")
        unit_used = st.session_state.get('forecast_unit_used', 'Days')
        periods_used = st.session_state.get('forecast_periods_used', 10)
        st.subheader(f"📊 Product-Wise Forecast Table ({periods_used} {unit_used})")
        
        forecast_df = st.session_state['forecast_table']
        avg_col = "Avg daily QTY" if unit_used == "Days" else "Avg weekly QTY"
        date_cols = [c for c in forecast_df.columns if c not in ['ITEMID', avg_col]]
        
        # Table Styling Logic
        def style_cells(val):
            try:
                val = float(val)
                if val >= high_thresh:
                    return 'background-color: #EAF3DE; color: #27500A; font-weight: 500'
                elif val >= med_thresh:
                    return 'background-color: #FAEEDA; color: #633806; font-weight: 500'
                else:
                    return 'background-color: #FCEBEB; color: #791F1F'
            except ValueError:
                return ''
                
        # Format values to integers, style cells, and display
        styled_df = forecast_df.style.map(style_cells, subset=date_cols)
        styled_df = styled_df.format("{:.0f}", subset=date_cols)
        
        st.dataframe(styled_df, use_container_width=True, height=400)
        
        # Export Excel Button
        excel_data = export_to_excel_styled(forecast_df, date_cols, med_thresh, high_thresh)
        st.download_button(
            label="📥 Download styled Excel Report",
            data=excel_data,
            file_name=f"prophet_{periods_used}{unit_used.lower()}_product_forecast.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # 5. Product-Level Deep Dive
        st.markdown("---")
        st.subheader("🔍 Single Product Deep Dive")
        selected_product = st.selectbox(
            "Select a product to inspect history, forecast, and seasonal patterns",
            options=unique_items
        )
        
        if selected_product:
            with st.spinner(f"Analyzing {selected_product}..."):
                # Get historical series
                df_series = prepare_product_series(df_cleaned, selected_product)
                
                # Get selected settings from session state or current config
                unit_used = st.session_state.get('forecast_unit_used', forecast_unit)
                periods_used = st.session_state.get('forecast_periods_used', forecast_periods)
                days_to_predict = periods_used if unit_used == "Days" else periods_used * 7

                # Fit and predict single product
                model, forecast_future = forecast_single_product(df_cleaned, selected_product, forecast_days=days_to_predict)
                
                # Retrieve full history + forecast to plot components
                future_all = model.make_future_dataframe(periods=days_to_predict, freq='D')
                forecast_full = model.predict(future_all)
                
                # Plot 1: History vs Forecast
                col_left, col_right = st.columns(2)
                
                with col_left:
                    st.write(f"### Historical vs Forecasted Sales ({selected_product})")
                    fig1, ax1 = plt.subplots(figsize=(10, 5))
                    ax1.plot(df_series['ds'][-120:], df_series['y'][-120:], label="Actual (Last 120 Days)", color="#1F77B4", linewidth=2)
                    
                    # Highlight forecast dates
                    forecast_dates = forecast_full[forecast_full['ds'] > df_series['ds'].max()]
                    ax1.plot(forecast_dates['ds'], forecast_dates['yhat'], label=f"Forecast ({periods_used} {unit_used})", color="#FF7F0E", linestyle="--", marker="o", linewidth=2)
                    ax1.fill_between(
                        forecast_dates['ds'], 
                        forecast_dates['yhat_lower'].clip(lower=0), 
                        forecast_dates['yhat_upper'].clip(lower=0), 
                        color="#FF7F0E", 
                        alpha=0.2, 
                        label="95% Confidence Interval"
                    )
                    ax1.set_xlabel("Date", fontsize=11)
                    ax1.set_ylabel("Quantity Sold", fontsize=11)
                    ax1.legend(loc="upper left")
                    ax1.grid(True, linestyle="--", alpha=0.5)
                    st.pyplot(fig1)
                    plt.close(fig1)
                    
                with col_right:
                    st.write(f"### Forecast Numbers ({selected_product})")
                    # Display values in a neat table (group weekly if using Weeks)
                    if unit_used == "Weeks":
                        forecast_future_disp = forecast_future.copy().reset_index(drop=True)
                        weekly_rows = []
                        for i in range(len(forecast_future_disp) // 7):
                            week_df = forecast_future_disp.loc[i*7 : (i+1)*7 - 1]
                            week_row = {
                                'Period': f"Week {i+1}",
                                'Expected Sale QTY': week_df['yhat'].sum(),
                                'Min QTY Bound': week_df['yhat_lower'].sum(),
                                'Max QTY Bound': week_df['yhat_upper'].sum()
                            }
                            weekly_rows.append(week_row)
                        forecast_future_disp = pd.DataFrame(weekly_rows)
                    else:
                        forecast_future_disp = forecast_future.copy()
                        forecast_future_disp['ds'] = forecast_future_disp['ds'].dt.strftime('%b %d, %Y')
                        forecast_future_disp.columns = ['Period', 'Expected Sale QTY', 'Min QTY Bound', 'Max QTY Bound']
                    st.table(forecast_future_disp)
                    
                # Plot 2: Seasonality Components
                st.write(f"### Prophet Seasonal & Holiday Components ({selected_product})")
                st.info("These plots explain the decomposition of the sales signal: general growth trend, public holiday influences, weekly patterns (sales peaks/troughs), and annual patterns.")
                fig2 = model.plot_components(forecast_full)
                st.pyplot(fig2)
                plt.close(fig2)

else:
    # Landing page layout when no file is uploaded
    st.info("👈 Please upload an Excel or CSV sales transaction file in the sidebar to begin.")
    
    st.markdown("""
    ### Expected Input File Format
    The file should contain transaction records with the following columns:
    
    | Column Name | Type | Description |
    | :--- | :--- | :--- |
    | **SlNo** | Integer | Invoice/transaction identifier |
    | **INVOICEDATE** | Date | Date of the sale (e.g. `YYYY-MM-DD`) |
    | **ITEMID** | String | Product/SKU identifier |
    | **QTY** | Integer | Quantity sold (negative values are cleaned automatically) |
    
    ### Features of the App:
    1. **Data Preprocessing**: Filters out negative quantities, handles sparse date gaps, and trims date ranges.
    2. **Product-wise Modeling**: Runs parallelized Prophet forecasting models for each product.
    3. **Uncertainty Bounds**: Calculates lower and upper bounds for each predicted day.
    4. **Seasonal Decomposition**: Inspects the growth trend, weekly cycles, and yearly cycles for each item.
    5. **Public Holiday Adjustment**: Incorporates Indian public holiday effects on demand.
    """)
